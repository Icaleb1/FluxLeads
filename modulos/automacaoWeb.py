import os
import re
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from urllib.parse import quote
from time import sleep
import random
import logging
from openpyxl.styles import PatternFill
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from modulos.interface import reiniciar_envios
from modulos.manipularArquivos import *


def normalizar_telefone(telefone):
    telefone_str = str(telefone)

    telefone_normalizado = re.sub(r'\D', '', telefone_str)

    if len(telefone_normalizado) == 11:  
        telefone_normalizado = '55' + telefone_normalizado
    elif len(telefone_normalizado) == 10:  
        telefone_normalizado = '55' + telefone_normalizado

    return telefone_normalizado

def enviar_anexo(navegador, telefone, anexo):
    intervaloAleatorio = random.uniform(20, 30)
    try:
        link_mensagem_whats = f'https://web.whatsapp.com/send?phone={telefone}'
        navegador.get(link_mensagem_whats)

        WebDriverWait(navegador, 25).until(
            EC.presence_of_element_located((By.ID, 'side'))   
        )
    
        anexar_icone = WebDriverWait(navegador, 25).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@title='Anexar']"))
        )
        anexar_icone.click()

        anexar_documento = navegador.find_element(By.XPATH, "//input[@accept='*']")
        anexar_documento.send_keys(anexo)

        sleep(intervaloAleatorio)
        botao_enviar = WebDriverWait(navegador, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
        )
        botao_enviar.click()

        sleep(intervaloAleatorio)
        return True

    except (NoSuchElementException, Exception) as e:
        logging.error(f"Erro ao enviar anexo para {telefone}: {e}")
        return False


def inicializar_navegador():
    options = webdriver.ChromeOptions()
    #options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('start-maximized')
    options.add_argument('window-size=800,600') 
    options.add_argument('window-position=-1000,-1000') 
    options.add_argument("user-data-dir=/path/to/your/custom/profile")
    navegador = webdriver.Chrome(options=options)
    navegador.get("https://web.whatsapp.com/")
    while len(navegador.find_elements(by='id', value='side')) < 1:
        sleep(15)
    return navegador

def processar_clientes(navegador, mensagens, pagina_clientes, workbook, file_path, anexo):

    if(verificar_arquivo_aberto(file_path)):
        messagebox.showwarning('Sistema','O excel não pode estar aberto!')
        return 

    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    intervaloAleatorio = random.uniform(20, 35)

    coluna_envio = encontrar_coluna_enviados("Enviado", pagina_clientes)
    if coluna_envio is None:
        coluna_envio = adicionar_coluna_envio(pagina_clientes)

    reiniciar_envios(pagina_clientes, coluna_envio)
    workbook.save(file_path)

    contador_mensagens = 0
    contador_erros_consecutivos = 0

    for linha in pagina_clientes.iter_rows(min_row=2):

        if len(linha) < 3:
            logging.error(f"Linha com dados insuficientes: {linha}")
            continue

        column_tel = encontrar_coluna_telefones(pagina_clientes)
        column_nome = encontrar_coluna_nomes(pagina_clientes)

        nome = linha[column_nome].value
        telefone = linha[column_tel].value
        

        status_envio = linha[coluna_envio - 1].value 
        if status_envio == "Sucesso":
            continue 

        if nome is None or telefone is None:
            continue

        primeira_mensagem = f"Olá, {nome}\n\n{mensagens[0]}"

        for i, mensagem in enumerate(mensagens): 
            mensagem_atual = primeira_mensagem if i == 0 else mensagem
            telefone_normalizado = normalizar_telefone(telefone)

            try:
                link_mensagem_whats = f'https://web.whatsapp.com/send?phone={telefone_normalizado}&text={quote(mensagem_atual)}'
                navegador.get(link_mensagem_whats)

                while len(navegador.find_elements(by='id', value='side')) < 1:
                    sleep(intervaloAleatorio)

                
                navegador.find_element(by='xpath', value='//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div[1]/p').send_keys(Keys.ENTER)
                sleep(intervaloAleatorio)

                navegador.switch_to.window(navegador.window_handles[0])

                sleep(intervaloAleatorio)

                if anexo:
                    envio_anexo_sucesso = enviar_anexo(navegador, telefone_normalizado, anexo)
                    if not envio_anexo_sucesso:
                        raise Exception(f"Falha ao enviar o anexo para o telefone {telefone_normalizado}")
                    
                linha[coluna_envio - 1].value = 'Sucesso'
                for cell in linha:
                    cell.fill = green_fill
                workbook.save(file_path)

                contador_erros_consecutivos = 0
                contador_mensagens += 1

                if contador_mensagens >= 50:
                    navegador.quit()
                    navegador = inicializar_navegador()
                    contador_mensagens = 0

            except Exception as e:

                logging.error(f"Erro ao processar cliente {nome} ({telefone}): {e}")
                
                linha[coluna_envio - 1].value = 'Falha'
                for cell in linha:
                    cell.fill = red_fill
                workbook.save(file_path)

                contador_erros_consecutivos += 1  

                if contador_erros_consecutivos >= 10:
                    logging.error("Muitos erros consecutivos. O programa será encerrado.")
                    messagebox.showinfo("Aviso","Muitos erros ocorreram. Por favor, reinicie o programa.")
                    navegador.quit()  
                    exit(1)  

    try:
        workbook.save(file_path)
    except Exception as e:
        logging.error(f"Erro ao salvar o arquivo: {e}")




      