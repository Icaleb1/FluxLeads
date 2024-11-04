import os
import re
import sys
from openpyxl.utils import get_column_letter


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath('.')
    return os.path.join(base_path, relative_path)

def encontrar_coluna_enviados(cabecalho, pagina_clientes):
    for col_idx, cell in enumerate(pagina_clientes[1]):  
        if cell.value == cabecalho:
            return col_idx + 1
    return None

def encontrar_coluna_telefones(pagina_clientes):
    palavra_chave = r'tel\.?|telefone|telefones'
    for col_idx, cell in enumerate(pagina_clientes[1]):  
        if cell.value and re.search(palavra_chave, cell.value, re.IGNORECASE):
            return col_idx  
    return None

def encontrar_coluna_nomes(pagina_clientes):
    palavra_chave = r'nome\.?|client\.?'
    for col_idx, cell in enumerate(pagina_clientes[1]):  
        if cell.value and re.search(palavra_chave, cell.value, re.IGNORECASE):
            return col_idx   
    return None

def adicionar_coluna_envio(pagina_clientes):
    nova_coluna_index = pagina_clientes.max_column + 1
    coluna_letra = get_column_letter(nova_coluna_index)
    pagina_clientes[f'{coluna_letra}1'].value = 'Enviado'  
    return nova_coluna_index

def resetar_status_envio(pagina_clientes, coluna_envio):
    coluna_envio = encontrar_coluna_enviados('Enviado', pagina_clientes)
    if coluna_envio is not None:
        for linha in pagina_clientes.iter_rows(min_row=2):
            linha[coluna_envio - 1].value = None


def verificar_numeros_enviados(pagina_clientes, coluna_envio):
    if coluna_envio is None:
       coluna_envio = adicionar_coluna_envio(pagina_clientes)

    ja_enviados = False

    for linha in pagina_clientes.iter_rows(min_row=2):
        if linha[coluna_envio - 1].value is not None: 
            ja_enviados = True
            break  

    return ja_enviados, None  
